import pandas as pd
import numpy as np
from sqlalchemy import create_engine
import onnxruntime as rt
import openpyxl
from openpyxl.styles import Font, PatternFill
from openpyxl.utils import get_column_letter

# SQL Bağlantısı
engine = create_engine('mssql+pyodbc://usesen:usesen@DESKTOP-6QR83E3\\UGURMSSQL/FSM_Tickets?driver=SQL+Server')

# Mevcut veriyi çek
query = """
SELECT 
    Bolge,
    Teknisyen,
    MONTH(Olusturma_Tarihi) as Ay,
    COUNT(*) as Ticket_Sayisi
FROM Tickets
WHERE 
    YEAR(Olusturma_Tarihi) = 2024
    AND MONTH(Olusturma_Tarihi) <= 3
GROUP BY 
    Bolge,
    Teknisyen,
    MONTH(Olusturma_Tarihi)
ORDER BY 
    Bolge,
    Teknisyen,
    Ay
"""
mevsimsel_faktor = {
    'Ankara': {
        4: 1.0,    # Normal
        5: 1.0,    # Normal
        6: 0.9,    # Düşüş başlıyor (tatil)
        7: 0.7,    # Keskin düşüş (tatil dönemi)
        8: 0.6,    # En düşük (tatil dönemi)
        9: 1.2,    # Yüksek (dönüş dönemi)
        10: 1.1,   # Hala yüksek
        11: 1.0,   # Normal
        12: 0.9    # Yılsonu
    },
    'İstanbul-Avrupa': {
        4: 1.0,    # Normal
        5: 1.0,    # Normal
        6: 0.9,    # Düşüş başlıyor
        7: 0.7,    # Keskin düşüş (tatil)
        8: 0.6,    # En düşük (tatil)
        9: 1.2,    # Yüksek (dönüş)
        10: 1.1,   # Hala yüksek
        11: 1.0,   # Normal
        12: 0.9    # Yılsonu
    },
    'İstanbul-Anadolu': {
        4: 1.0,    # Normal
        5: 1.0,    # Normal
        6: 0.9,    # Düşüş başlıyor
        7: 0.7,    # Keskin düşüş (tatil)
        8: 0.6,    # En düşük (tatil)
        9: 1.2,    # Yüksek (dönüş)
        10: 1.1,   # Hala yüksek
        11: 1.0,   # Normal
        12: 0.9    # Yılsonu
    },
    'İzmir': {
        4: 1.0,    # Normal
        5: 1.1,    # Artış
        6: 1.2,    # Yüksek (yazlıklar açılıyor)
        7: 1.3,    # En yüksek (tam sezon)
        8: 1.3,    # En yüksek (tam sezon)
        9: 1.1,    # Hala yüksek
        10: 1.0,   # Normal
        11: 0.9,   # Düşüş
        12: 0.9    # Düşük
    },
    'Bursa': {
        4: 1.0,    # Normal
        5: 1.0,    # Normal
        6: 0.9,    # Hafif düşüş
        7: 0.8,    # Düşüş (tatil ama İst/Ank kadar değil)
        8: 0.8,    # Düşük
        9: 1.2,    # Yüksek (dönüş)
        10: 1.1,   # Hala yüksek
        11: 1.0,   # Normal
        12: 0.9    # Yılsonu
    },
    'Antalya': {
        4: 1.0,    # Normal
        5: 1.1,    # Artış başlıyor
        6: 1.2,    # Yüksek (tatilciler)
        7: 1.3,    # En yüksek (tatil sezonu)
        8: 1.3,    # En yüksek (tatil sezonu)
        9: 1.1,    # Hala yüksek
        10: 1.0,   # Normal
        11: 0.9,   # Düşüş
        12: 0.9    # Düşük
    }
}

df = pd.read_sql(query, engine)

# Feature'ları hazırla
feature_df = pd.get_dummies(df[['Bolge', 'Teknisyen', 'Ay']])
feature_columns = feature_df.columns

# ONNX modelini yükle
sess = rt.InferenceSession("D:/PhytonEgitimleri/ticket_tahmin_modeli.onnx")
input_name = sess.get_inputs()[0].name
label_name = sess.get_outputs()[0].name

# Excel raporu oluştur
wb = openpyxl.Workbook()
wb.remove(wb.active)

# Her bölge için ayrı sheet
for bolge in df['Bolge'].unique():
    ws = wb.create_sheet(bolge)
    
    # Başlıklar
    headers = ['Teknisyen'] + [f'2024-{ay:02d}' for ay in range(4, 13)] + ['TOPLAM']
    ws.append(headers)
    
    # Bölge teknisyenleri
    bolge_data = df[df['Bolge'] == bolge]
    teknisyenler = bolge_data['Teknisyen'].unique()
    
    # Bölge toplam satırı
    bolge_row = [bolge] 
    bolge_toplam = 0
    
    # İş günleri
    is_gunleri = {4: 22, 5: 22, 6: 20, 7: 23, 8: 22, 9: 21, 10: 22, 11: 21, 12: 22}
    
    # Teknisyen tahminleri
    teknisyen_tahminler = {}
    
    for teknisyen in teknisyenler:
        teknisyen_tahminler[teknisyen] = {}
        
        for ay in range(4, 13):
            # Tahmin için veri hazırla
            tahmin_data = pd.DataFrame({
                'Bolge': [bolge],
                'Teknisyen': [teknisyen],
                'Ay': [ay]
            })
            feature_df = pd.get_dummies(df[['Bolge', 'Teknisyen', 'Ay']])
            feature_columns = feature_df.columns  # Tüm feature'ların sırası burada saklanıyor

            # One-hot encoding
            tahmin_features = pd.get_dummies(tahmin_data[['Bolge', 'Teknisyen', 'Ay']])
            
            # Eksik kolonları ekle
            for col in feature_columns:
                if col not in tahmin_features.columns:
                    tahmin_features[col] = 0
            
            # Feature sırasını aynı yap
            tahmin_features = tahmin_features[feature_columns]
            
            # ONNX ile tahmin yap
            X_test = tahmin_features.values.astype(np.float32)
            pred = sess.run([label_name], {input_name: X_test})[0][0]
            
            # Mevsimsel faktör ve iş günü sayısına göre ayarla
            pred = pred * mevsimsel_faktor[bolge][ay] * (is_gunleri[ay] / 22)
            
            teknisyen_tahminler[teknisyen][ay] = int(pred)
    
    # Bölge toplamları
    for ay in range(4, 13):
        ay_toplam = sum(teknisyen_tahminler[tek][ay] for tek in teknisyenler)
        bolge_row.append(ay_toplam)
        bolge_toplam += ay_toplam
    
    bolge_row.append(bolge_toplam)
    ws.append(bolge_row)
    
    # Boş satır
    ws.append([])
    
    # Teknisyen satırları
    for teknisyen in teknisyenler:
        tek_row = [teknisyen]
        tek_toplam = 0
        
        for ay in range(4, 13):
            tahmin = teknisyen_tahminler[teknisyen][ay]
            tek_row.append(tahmin)
            tek_toplam += tahmin
        
        tek_row.append(tek_toplam)
        ws.append(tek_row)
    
    # Formatlama
    header_fill = PatternFill(start_color="CCE5FF", end_color="CCE5FF", fill_type="solid")
    
    for row in ws.iter_rows():
        for cell in row:
            if isinstance(cell.value, (int, float)):
                cell.number_format = '#,##0'
            if cell.row == 1:
                cell.font = Font(bold=True)
                cell.fill = header_fill
            if cell.row == 2:
                cell.font = Font(bold=True)
    
    # Sütun genişlikleri
    ws.column_dimensions['A'].width = 20
    for col in range(2, len(headers) + 1):
        ws.column_dimensions[get_column_letter(col)].width = 12

# Kaydet
try:
    wb.save('D:/PhytonEgitimleri/ticket_tahminleri.xlsx')
    print("Tahminler başarıyla Excel'e kaydedildi!")
except PermissionError:
    print("Excel dosyası açık! Lütfen kapatıp tekrar deneyin.")
    wb.save('D:/PhytonEgitimleri/ticket_tahminleri_new.xlsx')
